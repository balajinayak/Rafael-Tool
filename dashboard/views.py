from django.shortcuts import render, redirect
from .models import SIPlan, AllSerial, SerialNumber, Mapping, NCRTracker, SIShipped, AllSIShipped, POTracker, DefectiveLibrary
from .forms import SiPlanForm, SerialNumberForm, MappingForm, NCRTrackerForm, POTrackerForm, NCRTrackerUpdateForm, SIShippedAddForm, DefectiveLibraryForm
from django.db.models import Count
from django.http import JsonResponse
from django.contrib import messages
from .models import *
from .forms import *
from .decorators import *



def si_plan_view(request):
    si_plans = SIPlan.objects.order_by('product', '-id').distinct('product')
    form = SiPlanForm()
    if request.method == 'POST':
        form = SiPlanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('si-plan')
            messages.success(request, 'Data added successfuly')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context = {
        'si_plans' : si_plans,
        'form' : form,
    }
    return render(request, 'dashboard/si_plan.html', context)




def update_si_plan_view(request, pk):
    si_plan = SIPlan.objects.get(id=pk)
    form = SiPlanForm(instance=si_plan)
    if request.method == 'POST':
        form = SiPlanForm(instance=si_plan, data=request.POST)
        if form.is_valid():
            fm = form.save(commit=False)
            SIPlan.objects.create(
                date=fm.date,
                project=fm.project,
                product=fm.product,
                cdlm_part_number=fm.cdlm_part_number,
                description=fm.description,
                po=fm.po,
                line=fm.line,
                offering_quantity=fm.offering_quantity,
                total_quantity=fm.total_quantity,
                si_level=fm.si_level,
                mrb_numbers=fm.mrb_numbers,
                remarks=fm.remarks,
            )
            messages.success(request, "Data updated successfully")
            return redirect('si-plan')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context = {
        'form' : form,
    }
    return render(request, 'dashboard/update_siplan.html', context)



def serial_number_view(request):
    ser_numbers = SerialNumber.objects.order_by('id')
    form = SerialNumberForm()
    if request.method == 'POST':
        form = SerialNumberForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Data added sucessfully")
            return redirect('serial-number')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    
    context = {
        'form' : form,
        'ser_numbers' : ser_numbers,
    }
    return render(request, 'dashboard/serial_number_verification.html', context)




def get_serial_numbers(request):
    data1 = AllSIShipped.objects.all()
    data2 = SerialNumber.objects.all()  # Replace with your query
    data1_list = [{"serial": serial.strip()} for item in data1 for serial in item.serial_number.split('\r\n')]
    data2_list = [{"serial": serial.strip()} for item in data2 for serial in item.serial_number.split('\r\n')]
    data_list = data1_list + data2_list
    return JsonResponse(data_list, safe=False)

    

import pandas as pd
from django.http import HttpResponse
import xlsxwriter
from datetime import datetime

def si_plan_export(request):
    if request.method == "POST":
        date_select = request.POST.get("date")
        data = SIPlan.objects.filter(date__lte=date_select).values()
        if data:

            df = pd.DataFrame(data)

            fields_to_include = ['date', 'project', 'product', 'cdlm', 'description', 'po', 'line', 'offering_quantity', 'total_quantity', 'si_level', 'mrb_numbers', 'remarks']

            df = pd.DataFrame(data, columns=fields_to_include)

            df.columns = df.columns.str.upper()

            current_date = datetime.now().strftime('%Y-%m-%d')
            df.insert(0, 'Sl.No', range(1, len(df) + 1))
            

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="SIPlans_{current_date}.xlsx"'

            with pd.ExcelWriter(response, engine='xlsxwriter', mode='xlsx') as writer:
                df.to_excel(writer, sheet_name='SIPlans', index=False)

            return response

        else:
            print("no data")
            return redirect('si-plan')
            

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import SerialNumber
from openpyxl.utils import get_column_letter


def serial_number_export(request):
    try:
        if request.method == "POST":
            date_from = request.POST.get("datefrom")
            date_to = request.POST.get("dateto")
            data = SerialNumber.objects.filter(date__gte=date_from, date__lte=date_to).order_by('id')

            if data:
                workbook = openpyxl.Workbook()
                sheet = workbook.active

                header_labels = ['Sl No', 'Project', 'PO Number', 'NCR Number', 'Date', 'Customer P/N', 'Cyient P/N', 'Quantity', 'Serial Numbers']

                for i, label in enumerate(header_labels, start=1):
                    sheet.cell(row=i, column=1, value=label).font = Font(bold=True)

                for index, serial_number in enumerate(data, start=2):
                    col = get_column_letter(index)
                    sheet[f'{col}1'] = index - 1
                    sheet[f'{col}2'] = serial_number.project
                    sheet[f'{col}3'] = serial_number.po_number
                    sheet[f'{col}4'] = serial_number.ncr_number
                    sheet[f'{col}5'] = serial_number.date
                    sheet[f'{col}6'] = serial_number.customer_part_number
                    sheet[f'{col}7'] = serial_number.cyient_part_number
                    sheet[f'{col}8'] = serial_number.quantity

                    serial_numbers = serial_number.serial_number.splitlines()  # Split without including newline characters
                    for ind, sn in enumerate(serial_numbers):
                        sheet[f'{col}{ind+9}'] = sn

                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="SerialNumbers_verification.xlsx"'

                workbook.save(response)

                return response
            else:
                messages.error(request, "No Data")
                return redirect('si-plan')
    except Exception as e:
        messages.error(request, f"An error occurred: {e}")
        return redirect('si-plan')





def serial_mapping_view(request):

    data = Mapping.objects.order_by('-date')

    form = MappingForm()
    if request.method == 'POST':
        form = MappingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Data added successfully")
            return redirect('serial-mapping')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    
    context={
        'form' : form,
        'data' : data,
    }
    return render(request, 'dashboard/serial_number_mapping.html', context)



from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def ncr_tracking_view(request):
    data = Ncrdump.objects.order_by('-timestamp')

    context = {
        'data': data,
    }
    return render(request, 'dashboard/ncr_tracker.html', context)



def add_ncr_tracking_view(request):
    form = NCRTrackerForm()
    if request.method == 'POST':
        form = NCRTrackerForm(request.POST, request.FILES)
        if form.is_valid():
            fm = form.save(commit=False)
            fm.remaining = fm.ncr_quantity
            fm.save()
            messages.success(request, "Data added successfully")
            return redirect('ncr-tracking')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
    }
    return render(request, 'dashboard/add_ncr_tracker.html', context)



from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

def ncr_dump_view(request):
    data = NCRTracker.objects.order_by('internal_fg_partnumber', '-timestamp').distinct('internal_fg_partnumber')
    form = NcrdumpForm()
    
    if request.method == 'POST':
        form = NcrdumpForm(request.POST, request.FILES)
        if form.is_valid():
            fm = form.save(commit=False)
            fm.save()
            messages.success(request, "Data added successfully")
            return redirect('ncr-tracking')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    
    paginator = Paginator(data, 10)  # Show 10 data per page

    page_number = request.GET.get('page')
    try:
        paginated_data = paginator.page(page_number)
    except PageNotAnInteger:
        paginated_data = paginator.page(1)
    except EmptyPage:
        paginated_data = paginator.page(paginator.num_pages)

    context = {
        'form': form,
        'data': paginated_data,
    }
    return render(request, 'dashboard/ncr_dump.html', context)



from django.db import transaction

def upload_excel_view(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return redirect('ncr_dump')
        try:
            df = pd.read_excel(excel_file)

            # Validate required columns
            required_columns = ['project', 'date', 'ncrno', 'ccaapp', 'infgpn', 'ponum',
                                'ncrdetails', 'proslno', 'qa', 'ncrqty', 'poqty', 'shipqty',
                                'openqty', 'ncrstatus', 'approvalstatus', 'remarks']
            if not all(col in df.columns for col in required_columns):
                messages.error(request, 'Missing required columns in Excel file.')
                return redirect('ncr_dump')

            with transaction.atomic():
                for index, row in df.iterrows():
                    # Replace NaN values with None
                    row = row.where(pd.notnull(row), None)

                    Ncrdump.objects.create(
                        project=row['project'],
                        date=row['date'],
                        ncrno=row['ncrno'],
                        ccaapp=row['ccaapp'],
                        infgpn=row['infgpn'],
                        ponum=row['ponum'],
                        ncrdetails=row['ncrdetails'],
                        proslno=row['proslno'],
                        qa=row['qa'],
                        ncrqty=row['ncrqty'],
                        poqty=row['poqty'],
                        shipqty=row['shipqty'],
                        openqty=row['openqty'],
                        ncrstatus=row['ncrstatus'],
                        approvalstatus=row['approvalstatus'],
                        remarks=row['remarks'],
                    )

            messages.success(request, 'Data uploaded successfully.')
            return redirect('ncr_dump')
        except Exception as e:
            messages.error(request, f'Error uploading data: {str(e)}')
            return redirect('ncr_dump')
    return render(request, 'update/ncr_dump.html')


def update_ncr_tracking_view(request, pk):
    data = NCRTracker.objects.get(id=pk)
    form = NCRTrackerUpdateForm()
    if request.method == 'POST':
        form = NCRTrackerUpdateForm(instance=data, data=request.POST)
        if form.is_valid():
            fm = form.save(commit=False)
            add_quantity = form.cleaned_data.get('add_quantity')
            if add_quantity <= 0:
                messages.error(request, "Value should not be lessethan 1")
            else:
                if add_quantity > data.remaining:
                    messages.error(request, "NCR quantity already consumed")
                else:
                    NCRTracker.objects.create(
                        project=data.project,
                        serial_number=data.serial_number,
                        ncr_date=data.ncr_date,
                        ncr_number=data.ncr_number,
                        customer_part_number=data.customer_part_number,
                        internal_fg_partnumber=data.internal_fg_partnumber,
                        po_number=data.po_number,
                        details=data.details,
                        ncr_quantity=data.ncr_quantity,
                        consumed = data.consumed+add_quantity,
                        remaining = data.remaining - add_quantity,
                        add_quantity=add_quantity,
                        ncr_status=data.ncr_status,
                        approval_status=data.approval_status,
                        approval_Date=data.approval_Date,
                        document=data.document,
                        site=data.site,
                        customer=data.customer,
                        product_description=data.product_description,
                        product_part_number=data.product_part_number,
                        issue_reported_stage=data.issue_reported_stage,
                        po_details=data.po_details,
                        product_rev=data.product_rev,
                        cyient_fg_number=data.cyient_fg_number,
                        issued_by=data.issued_by,
                        issued_to=data.issued_to,
                        lot=data.lot,
                        lot_and_defect_status = data.lot_and_defect_status,
                        image1=data.image1,
                        image2=data.image2,
                        circuit_reference_or_ref_designator=data.circuit_reference_or_ref_designator,
                        number_of_defects=data.number_of_defects,
                        containment_action=data.containment_action,
                        containment_resp=data.containment_resp,
                        containment_date=data.containment_date,
                        root_cause_analysis=data.root_cause_analysis,
                        why1=data.why1,
                        why1_man=data.why1_man,
                        why1_method=data.why1_method,
                        why1_material=data.why1_material,
                        why1_machine=data.why1_machine,
                        why2=data.why2,
                        why2_man=data.why2_man,
                        why2_method=data.why2_method,
                        why2_material=data.why2_material,
                        why2_machine=data.why2_machine,
                        why3=data.why3,
                        why3_man=data.why3_man,
                        why3_method=data.why3_method,
                        why3_material=data.why3_material,
                        why3_machine=data.why3_machine,
                        why4=data.why4,
                        why4_man=data.why4_man,
                        why4_method=data.why4_method,
                        why4_material=data.why4_material,
                        why4_machine=data.why4_machine,
                        why5=data.why5,
                        why5_man=data.why5_man,
                        why5_method=data.why5_method,
                        why5_material=data.why5_material,
                        why5_machine=data.why5_machine,
                        how=data.how,
                        corrective_action=data.corrective_action,
                        corrective_resp=data.corrective_resp,
                        corrective_date=data.corrective_date,
                        preventive_action=data.preventive_action,
                        preventive_resp=data.preventive_resp,
                        preventive_date=data.preventive_date,
                        wo_no=data.wo_no,
                        batch_no=data.batch_no,
                        verification_details=data.verification_details,
                        verification_status=data.verification_status,
                        verification_resp=data.verification_resp,
                        varifiction_date=data.varifiction_date,
                    )
                    messages.success(request, "Data updated successfully")
                    return redirect('ncr-tracking')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
        'data' : data,
    }
    return render(request, 'dashboard/update_ncr_tracker.html', context)

from django.shortcuts import render, get_object_or_404, redirect
from .models import Ncrdump
from .forms import NCRdumpupdateForm

def update_ncrdump_view(request, pk):
    ncrdump = get_object_or_404(Ncrdump, pk=pk)
    form = NCRdumpupdateForm(request.POST or None, instance=ncrdump)
    
    if form.is_valid():
        add_qty = form.cleaned_data['add_qty']
        if add_qty < 0:
            form.add_error('add_qty', 'Add quantity must be greater than 0')
        elif add_qty > ncrdump.openqty:
            form.add_error('add_qty', 'Add quantity should not be greater than OPEN qty')
        elif add_qty > ncrdump.ncrqty:
            form.add_error('add_qty', 'Add quantity should not be greater than NCR qty')
        else:
            ncrdump.shipqty += add_qty
            ncrdump.openqty -= add_qty
            ncrdump.save()
            return redirect('ncr-tracking')
    
    return render(request, 'dashboard/update_ncrdump.html', {'form': form, 'ncrdump': ncrdump})




import os
import pandas as pd
from django.http import HttpResponse
from django.views import View
from .models import NCRTracker
from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import tempfile
from openpyxl.styles import Alignment

class ExportNCRDataToExcel(View):
    def get(self, request):
        try:
            # Fetch data from the NCRTracker model
            queryset = NCRTracker.objects.all()

            # Create an Excel workbook and add a worksheet
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="ncr_data.xlsx"'
            wb = Workbook()
            ws = wb.active

            # Add headers to the worksheet
            headers = [
                "Sl No",
                "Project",
                "Serial Number",
                "NCR Date",
                "NCR Number",
                "Customer Part Number",
                "Internal FG Part Number",
                "PO Number",
                "Details",
                "Quantity",
                "Remaining",
                "Consumed",
                "Added Quantity",
                "NCR Status",
                "Approval Status",
                "Approval Date",
                "Site",
                "Customer",
                "Product Description",
                "Issue Reported Stage",
                "PO Details",
                "Product Rev",
                "Issued by",
                "Issued to",
                "LOT",
                "LOT and Defect Status",
            ]
            ws.append(headers)

            # Create a temporary directory for storing images
            temp_dir = tempfile.mkdtemp()

            # Initialize row index for image placement
            img_row = 2  # Start from the second row (assuming the header is in the first row)

            # Set column widths to accommodate text
            ws.column_dimensions['A'].width = 10  # Adjust the width as needed
            ws.column_dimensions['B'].width = 20  # Adjust the width as needed
            ws.column_dimensions['C'].width = 20  # Adjust the width as needed
            ws.column_dimensions['D'].width = 15  # Adjust the width as needed
            ws.column_dimensions['E'].width = 20   # Width for the gap column
            ws.column_dimensions['F'].width = 20  # Adjust the width as needed
            ws.column_dimensions['G'].width = 20  # Adjust the width as needed
            ws.column_dimensions['H'].width = 20  # Adjust the width as needed
            ws.column_dimensions['I'].width = 20  # Adjust the width as needed
            ws.column_dimensions['J'].width = 20  # Adjust the width as needed

            # Add a column for the image
            ws.cell(row=1, column=27, value="Image")

            # Iterate through the queryset and add data to the worksheet
            for item in queryset:
                # Add the data to the same row
                row_data = [
                    item.id,
                    item.project,
                    item.serial_number,
                    item.ncr_date,
                    item.ncr_number,
                    item.customer_part_number,
                    item.internal_fg_partnumber,
                    item.po_number,
                    item.details,
                    item.ncr_quantity,
                    item.remaining,
                    item.consumed,
                    item.add_quantity,
                    item.ncr_status,
                    item.approval_status,
                    item.approval_Date,
                    item.site,
                    item.customer,
                    item.product_description,
                    item.issue_reported_stage,
                    item.po_details,
                    item.product_rev,
                    item.issued_by,
                    item.issued_to,
                    item.lot,
                    item.lot_and_defect_status,
                ]

                # Append a blank cell for the image
                row_data.append('')

                ws.append(row_data)

                # Add the image to the worksheet
                image_path = os.path.join(settings.MEDIA_ROOT, str(item.image))
                if os.path.exists(image_path):
                    img = Image.open(image_path)
                    img = img.resize((1000, 1000))  # Resize the image as needed
                    img_temp_path = os.path.join(temp_dir, f'temp_image_{img_row}.png')
                    img.save(img_temp_path)
                    img = ExcelImage(img_temp_path)

                    # Set the anchor point to position the image in the last column
                    img.anchor = f'AA{img_row}'  # Adjust the row as needed
                    img.width = 200  # Adjust the width of the image
                    img.height = 140  # Adjust the height of the image
                    ws.add_image(img)

                    # Set the row height to match the image height
                    img_height = img.height
                    ws.row_dimensions[img_row].height = img_height  # Adjust the row as needed

                    # Align text at the top of the cell
                    for cell in ws[f"A{img_row}:Z{img_row}"]:
                        for c in cell:
                            c.alignment = Alignment(vertical='top')

                    img_row += 1  # Move to the next row for the next data and image

            # Save the workbook to the response
            wb.save(response)

            # Close the workbook
            wb.close()

            # Clean up the temporary directory
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            os.rmdir(temp_dir)

            return response
        except Exception as e:
            # Handle any exceptions and return an error response
            return HttpResponse(f"An error occurred: {str(e)}", status=500)



class ExportNCRData(View):
    def get(self, request):
        try:
            # Fetch data from the NCRTracker model
            queryset = NCRTracker.objects.all()

            # Create an Excel workbook and add a worksheet
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="ncr_data.xlsx"'
            wb = Workbook()
            ws = wb.active

            # Add headers to the worksheet
            headers = [
                "Sl No",
                "Project",
                "Serial Number",
                "NCR Date",
                "NCR Number",
                "Customer Part Number",
                "Internal FG Part Number",
                "PO Number",
                "Details",
                "Quantity",
                "Remaining",
                "Consumed",
                "Added Quantity",
                "NCR Status",
                "Approval Status",
                "Approval Date",
            ]
            ws.append(headers)

            # Create a temporary directory for storing images
            temp_dir = tempfile.mkdtemp()

            # Set column widths to accommodate text
            ws.column_dimensions['A'].width = 10  # Adjust the width as needed
            ws.column_dimensions['B'].width = 20  # Adjust the width as needed
            ws.column_dimensions['C'].width = 20  # Adjust the width as needed
            ws.column_dimensions['D'].width = 15  # Adjust the width as needed
            ws.column_dimensions['E'].width = 20   # Width for the gap column
            ws.column_dimensions['F'].width = 20  # Adjust the width as needed
            ws.column_dimensions['G'].width = 20  # Adjust the width as needed
            ws.column_dimensions['H'].width = 20  # Adjust the width as needed
            ws.column_dimensions['I'].width = 20  # Adjust the width as needed
            ws.column_dimensions['J'].width = 20  # Adjust the width as needed

            # Iterate through the queryset and add data to the worksheet
            for item in queryset:
                # Add the data to the same row
                row_data = [
                    item.id,
                    item.project,
                    item.serial_number,
                    item.ncr_date,
                    item.ncr_number,
                    item.customer_part_number,
                    item.internal_fg_partnumber,
                    item.po_number,
                    item.details,
                    item.ncr_quantity,
                    item.remaining,
                    item.consumed,
                    item.add_quantity,
                    item.ncr_status,
                    item.approval_status,
                    item.approval_Date,
                ]
                ws.append(row_data)

            # Save the workbook to the response
            wb.save(response)

            # Close the workbook
            wb.close()

            # Clean up the temporary directory
            for filename in os.listdir(temp_dir):
                file_path = os.path.join(temp_dir, filename)
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            os.rmdir(temp_dir)

            return response
        except Exception as e:
            # Handle any exceptions and return an error response
            return HttpResponse(f"An error occurred: {str(e)}", status=500)




def po_tracking_view(request):
    data = POTracker.objects.order_by('rafael_part_number', '-id')

    context={
        'data' : data,
    }
    return render(request, 'dashboard/po_tracker.html', context)

    


from .forms import POTrackerForm
def add_po_tracking_view(request):
    form = POTrackerForm()
    if request.method == 'POST':
        form = POTrackerForm(request.POST, request.FILES)
        if form.is_valid():
            fm = form.save(commit=False)
            fm.remaining = fm.quantity
            fm.save()
            messages.success(request, "Data added successfully")
            return redirect('po-tracking')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
    }
    return render(request, 'dashboard/add_po_tracker.html', context)


from django.db import transaction

def upload_excel_po(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return redirect('po-tracking')
        try:
            df = pd.read_excel(excel_file)

            # Validate required columns
            required_columns = ['Project', 'Customer', 'Ship to', 'Rafael PN', 'Cyient PN', 'Description',
                                'RAFAEL PO Number', 'KRAS PO Number', 'PO Line', 'Remaining Qty', 'Unit Price in $']
            if not all(col in df.columns for col in required_columns):
                messages.error(request, 'Missing required columns in Excel file.')
                return redirect('po-tracking')

            with transaction.atomic():
                for index, row in df.iterrows():
                    # Replace NaN values with None
                    row = row.where(pd.notnull(row), None)

                    POTracker.objects.create(
                        project=row['Project'],
                        customer=row['Customer'],
                        ship_to=row['Ship to'],
                        rafael_part_number=row['Rafael PN'],
                        cyient_part_number=row['Cyient PN'],
                        description=row['Description'],
                        po_number=row['RAFAEL PO Number'],
                        kras_po_number=row['KRAS PO Number'],
                        poline=row['PO Line'],
                        quantity=row['Remaining Qty'],
                        unit_price=row['Unit Price in $'],
               
                    )

            messages.success(request, 'Data uploaded successfully.')
            return redirect('po-tracking')
        except Exception as e:
            messages.error(request, f'Error uploading data: {str(e)}')
            return redirect('po-tracking')
    return render(request, 'dashboard/po-tracking.html')


from .forms import POTrackerUpdateForm

def update_po_tracking_view(request, pk):
    data = POTracker.objects.get(id=pk)
    form = POTrackerUpdateForm()
    if request.method == 'POST':
        form = POTrackerUpdateForm(instance=data, data=request.POST)
        if form.is_valid():
            fm = form.save(commit=False)
            add_quantity = form.cleaned_data.get('add_quantity')
            if add_quantity <= 0:
                messages.error(request, "Value should not be lessethan 1")
            else:
                if add_quantity > data.remaining:
                    messages.error(request, "NCR quantity already consumed")
                else:
                    POTracker.objects.create(
                        project=data.project,
                        customer=data.customer,
                        ship_to=data.ship_to,
                        rafael_part_number=data.rafael_part_number,
                        cyient_part_number=data.cyient_part_number,
                        description=data.description,
                        po_number=data.po_number,
                        kras_po_number=data.kras_po_number,
                        quantity=data.quantity,
                        unit_price=data.unit_price,
                        consumed = data.consumed + add_quantity,
                        remaining = data.remaining - add_quantity,
                        add_quantity=add_quantity,
                    )
                    messages.success(request, "Data updated successfully")
                    return redirect('po-tracking')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
        'data' : data,
    }
    return render(request, 'dashboard/update_po_tracker.html', context)

from io import BytesIO
def po_tracker_export(request):
    data = POTracker.objects.all()

    workbook = Workbook()
    worksheet = workbook.active

    header = ["Sl No", "Project", "Customer", "Ship To", "Rafael Part Number", "Cyient Part Number", "Description", "Rafael PO Number","Kras PO Number", "Quantity","Unit Price", "Consumed", "Remaining", "Add Quantity", "Timestamp"]
    worksheet.append(header)

    for item in data:
        timestamp_naive = item.timestamp.replace(tzinfo=None)

        row = [item.id, item.project, item.customer, item.ship_to, item.rafael_part_number, item.cyient_part_number, item.description, item.po_number,item.kras_po_number, item.quantity,item.unit_price, item.consumed, item.remaining, item.add_quantity, timestamp_naive]
        worksheet.append(row)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="po_tracker.xlsx"'

    output = BytesIO()
    workbook.save(output)
    response.write(output.getvalue())

    return response



def defect_library_view(request):
    data = DefectiveLibrary.objects.order_by('-product_part_number')

    context={
        'data' : data,
    }
    return render(request, 'dashboard/defect_library.html', context)

@superuser_required
def add_defect_library_view(request):
    form = DefectiveLibraryForm()
    if request.method == 'POST':
        form = DefectiveLibraryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Data updated successfully")
            return redirect('defect-library')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
    }
    return render(request, 'dashboard/add_defect_library.html', context)


def update_defect_library_view(request, pk):
    data = DefectiveLibrary.objects.get(id=pk)
    form = DefectiveLibraryForm(instance=data)
    if request.method == 'POST':
        form = DefectiveLibraryForm(request.POST, request.FILES, instance=data)
        if form.is_valid():
            fm = form.save(commit=False)
            DefectiveLibrary.objects.create(
                customer=fm.customer,
                project=fm.project,
                product_part_number=fm.product_part_number,
                brief_defect_description=fm.brief_defect_description,
                customer_part_number=fm.customer_part_number,
                product_rev=fm.product_rev,
                product_description=fm.product_description,
                complaint_description=fm.complaint_description,
                source_of_complaint=fm.source_of_complaint,
                quantity_affected=fm.quantity_affected,
                defect_image_available=fm.defect_image_available,
                original_date=fm.original_date,
                defect_image=fm.defect_image,
                root_cause=fm.root_cause,
                ca_pa=fm.ca_pa,
                defect_cause=fm.defect_cause,
            )
            messages.success(request, "Data updated successfully")
            return redirect('defect-library')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form' : form,
    }
    return render(request, 'dashboard/add_defect_library.html', context)


def defect_library_details_view(request, pk):
    data = DefectiveLibrary.objects.get(id=pk)
    context={
        'data' : data,
    }
    return render(request, 'dashboard/defect_library_details.html', context)


from .resources import ExcelResources
from tablib import Dataset
from django.shortcuts import render, redirect
from django.contrib import messages
'''
def si_shipped_view(request):

    data = AllSIShipped.objects.order_by('-id')

    if request.method == 'POST':
        excel_resource = ExcelResources()
        dataset = Dataset()
        try:
            new_data = request.FILES['myfile']
            if not new_data.name.endswith('xlsx'):
                messages.error(request, 'Wrong format')
                return render(request, 'dashboard/si_shipped.html')
            
            # Associate the dataset with the excel_resource
            dataset.load(new_data.read(), format='xlsx', headers=True)
            dataset.append_col([None] * len(dataset), header='id')  # Add 'id' column
            
            for row in dataset:
                # Access elements using integer indices
                value = SIShipped(
                    sl_no=row[0],
                    invoice_date=row[1],  # Assuming 'Invoice Date' is the first column (index 0)
                    customer_part_number=row[2],  # Assuming 'Customer Part No' is the second column (index 1)
                    cdlm_part_number=row[3],  # Assuming 'CDLM Part No' is the third column (index 2)
                    invoice_no=row[4],  # Assuming 'Invoice No' is the fourth column (index 3)
                    po_number=row[5],  # Assuming 'PO No' is the fifth column (index 4)
                    customer=row[6],  # Assuming 'Customer' is the sixth column (index 5)
                    serial_number=row[7],  # Assuming 'Serial Number' is the seventh column (index 6)
                    quantity=row[8],  # Assuming 'Quantity' is the eighth column (index 7)
                    remarks=row[9]  # Assuming 'Remarks' is the ninth column (index 8)
                )
                try:
                    value.save()
                except Exception as e:
                    messages.error(request, f"Error saving data: {e}", extra_tags='danger')
            
            messages.success(request, "Excel data added successfully")
            return redirect('si-shipped')
        
        except Exception as e:
            messages.error(request, f"Error processing file: {e}", extra_tags='danger')
            return redirect('si-shipped')
    
    context = {
        'data' : data,
    }
    return render(request, 'dashboard/si_shipped.html', context)
'''

from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import SIShipped
import pandas as pd
from dateutil import parser

def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file, header=None)  # Read without header
            for index, row in df.iterrows():
                try:
                    sl_no = int(row[0]) if len(row) > 0 and str(row[0]).strip().isdigit() else None
                    invoice_date = parser.parse(str(row[1])).strftime('%Y-%m-%d') if len(row) > 1 else None
                    customer_part_number = row[2] if len(row) > 2 else None
                    cdlm_part_number = row[3] if len(row) > 3 else None
                    invoice_no = row[4] if len(row) > 4 else None
                    po_number = row[5] if len(row) > 5 else None
                    customer = row[6] if len(row) > 6 else None
                    serial_number = row[7] if len(row) > 7 else None
                    remarks = row[8] if len(row) > 8 else None
                    SIShipped.objects.create(
                        sl_no=sl_no,
                        invoice_date=invoice_date,
                        customer_part_number=customer_part_number,
                        cdlm_part_number=cdlm_part_number,
                        invoice_no=invoice_no,
                        po_number=po_number,
                        customer=customer,
                        serial_number=serial_number,
                        remarks=remarks
                    )
                except ValueError:
                    pass  # Skip row if sl_no is not a valid integer
            return render(request, 'dashboard/si_shipped.html')
    else:
        form = ExcelUploadForm()
    return render(request, 'dashboard/si_shipped.html', {'form': form})



    
def add_si_shipped(request):
    form = SIShippedAddForm()
    if request.method=="POST":
        form = SIShippedAddForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Data added successfully")
            return redirect('si-shipped')
        else:
            messages.error(request, 'Something went wrong', extra_tags='danger')
    context={
        'form':form,
    }
    return render(request, 'dashboard/add_si_shipped.html', context)



def delete_shipped_excel_view(request):
    data1 = SIShipped.objects.all()
    data2 = AllSIShipped.objects.all()
    for d1 in data1:
        d1.delete()
        
    for d2 in data2:
        d2.delete()

    return redirect('si-shipped')


# si shipped new

from django.shortcuts import render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import SIShipped
from .forms import SIShippedSearchForm
from django.db.models import Q

def search_sishipped(request):
    template_name = 'dashboard/si_shipped.html'
    search_form = SIShippedSearchForm()
    search_term = None

    if request.method == 'POST':
        search_form = SIShippedSearchForm(request.POST)
        if search_form.is_valid():
            search_term = search_form.cleaned_data.get('search_term')
            # Apply filtering based on the search term
            sishipped_data = SIShipped.objects.filter(
                Q(sl_no__icontains=search_term) |
                Q(invoice_date__icontains=search_term) |
                Q(customer_part_number__icontains=search_term) |
                Q(cdlm_part_number__icontains=search_term) |
                Q(invoice_no__icontains=search_term) |
                Q(po_number__icontains=search_term) |
                Q(customer__icontains=search_term) |
                Q(serial_number__icontains=search_term) |
                Q(quantity__icontains=search_term) |
                Q(remarks__icontains=search_term)
            ).order_by('-invoice_date')
        else:
            sishipped_data = SIShipped.objects.all().order_by('-invoice_date')
    else:
        sishipped_data = SIShipped.objects.all().order_by('-invoice_date')

    paginator = Paginator(sishipped_data, 10)  # Show 10 items per page
    page = request.GET.get('page')

    try:
        sishipped_data = paginator.page(page)
    except PageNotAnInteger:
        sishipped_data = paginator.page(1)
    except EmptyPage:
        sishipped_data = paginator.page(paginator.num_pages)

    return render(request, template_name, {'search_form': search_form, 'sishipped_data': sishipped_data, 'search_term': search_term})



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from django.http import HttpResponse
from openpyxl.styles import PatternFill

def demo_excel(request, pk):
    data = NCRTracker.objects.get(id=pk)
    workbook = Workbook()
    sheet = workbook.active

    # Background Color
    yellow_fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")

    for row in sheet.iter_rows(min_row=10, max_row=10, min_col=1, max_col=1):
        for cell in row:
            cell.fill = yellow_fill

    for row in sheet.iter_rows(min_row=13, max_row=13, min_col=1, max_col=20):
        for cell in row:
            cell.fill = yellow_fill

    for row in sheet.iter_rows(min_row=16, max_row=16, min_col=1, max_col=20):
        for cell in row:
            cell.fill = yellow_fill

    for row in sheet.iter_rows(min_row=25, max_row=25, min_col=1, max_col=20):
        for cell in row:
            cell.fill = yellow_fill

    for row in sheet.iter_rows(min_row=28, max_row=28, min_col=1, max_col=20):
        for cell in row:
            cell.fill = yellow_fill

    for row in sheet.iter_rows(min_row=31, max_row=31, min_col=1, max_col=20):
        for cell in row:
            cell.fill = yellow_fill

    # Width
    sheet.column_dimensions['A'].width=17

    # Merging
    sheet.merge_cells('A1:T1')

    sheet.merge_cells('A2:C2')
    sheet.merge_cells('D2:K2')
    sheet.merge_cells('L2:N2')
    sheet.merge_cells('O2:T2')

    sheet.merge_cells('A3:C3')
    sheet.merge_cells('D3:K3')
    sheet.merge_cells('L3:N3')
    sheet.merge_cells('O3:T3')


    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D4:K4')
    sheet.merge_cells('L4:N4')
    sheet.merge_cells('O4:T4')

    sheet.merge_cells('A5:C5')
    sheet.merge_cells('D5:F5')
    sheet.merge_cells('G5:H5')
    sheet.merge_cells('I5:K5')
    sheet.merge_cells('L5:N5')
    sheet.merge_cells('O5:T5')

    sheet.merge_cells('A6:C6')
    sheet.merge_cells('D6:K6')
    sheet.merge_cells('L6:N6')
    sheet.merge_cells('O6:T6')

    sheet.merge_cells('A7:C7')
    sheet.merge_cells('D7:K7')
    sheet.merge_cells('L7:N7')
    sheet.merge_cells('O7:T7')

    sheet.merge_cells('A8:C8')
    sheet.merge_cells('D8:K8')
    sheet.merge_cells('L8:N8')
    sheet.merge_cells('O8:T8')

    sheet.merge_cells('A9:T9')

    sheet.merge_cells('A10:T10')

    sheet.merge_cells('B11:M11')
    sheet.merge_cells('N11:Q11')
    sheet.merge_cells('R11:T11')

    sheet.merge_cells('B12:M12')
    sheet.merge_cells('N12:Q12')
    sheet.merge_cells('R12:T12')

    sheet.merge_cells('A13:M13')
    sheet.merge_cells('N13:Q13')
    sheet.merge_cells('R13:T13')

    sheet.merge_cells('A14:M14')
    sheet.merge_cells('N14:Q14')
    sheet.merge_cells('R14:T14')

    sheet.merge_cells('A15:M15')
    sheet.merge_cells('N15:Q15')
    sheet.merge_cells('R15:T15')

    sheet.merge_cells('A16:T16')

    sheet.merge_cells('B17:L17')
    sheet.merge_cells('M17:N17')
    sheet.merge_cells('O17:P17')
    sheet.merge_cells('Q17:R17')
    sheet.merge_cells('S17:T17')

    sheet.merge_cells('B18:L18')
    sheet.merge_cells('M18:N18')
    sheet.merge_cells('O18:P18')
    sheet.merge_cells('Q18:R18')
    sheet.merge_cells('S18:T18')

    sheet.merge_cells('B19:L19')
    sheet.merge_cells('M19:N19')
    sheet.merge_cells('O19:P19')
    sheet.merge_cells('Q19:R19')
    sheet.merge_cells('S19:T19')

    sheet.merge_cells('B20:L20')
    sheet.merge_cells('M20:N20')
    sheet.merge_cells('O20:P20')
    sheet.merge_cells('Q20:R20')
    sheet.merge_cells('S20:T20')

    sheet.merge_cells('B21:L21')
    sheet.merge_cells('M21:N21')
    sheet.merge_cells('O21:P21')
    sheet.merge_cells('Q21:R21')
    sheet.merge_cells('S21:T21')

    sheet.merge_cells('B22:L22')
    sheet.merge_cells('M22:N22')
    sheet.merge_cells('O22:P22')
    sheet.merge_cells('Q22:R22')
    sheet.merge_cells('S22:T22')

    sheet.merge_cells('B23:L23')
    sheet.merge_cells('M23:T23')

    sheet.merge_cells('A24:T24')

    sheet.merge_cells('A25:M25')
    sheet.merge_cells('N25:Q25')
    sheet.merge_cells('R25:T25')

    sheet.merge_cells('A26:M26')
    sheet.merge_cells('N26:Q26')
    sheet.merge_cells('R26:T26')

    sheet.merge_cells('A27:T27')

    sheet.merge_cells('A28:M28')
    sheet.merge_cells('N28:Q28')
    sheet.merge_cells('R28:T28')

    sheet.merge_cells('A29:M29')
    sheet.merge_cells('N29:Q29')
    sheet.merge_cells('R29:T29')

    sheet.merge_cells('A30:T30')

    sheet.merge_cells('A31:T31')

    sheet.merge_cells('A32:B32')
    sheet.merge_cells('C32:D32')
    sheet.merge_cells('E32:N32')
    sheet.merge_cells('O32:P32')
    sheet.merge_cells('Q32:R32')
    sheet.merge_cells('S32:T32')

    sheet.merge_cells('A33:B33')
    sheet.merge_cells('C33:D33')
    sheet.merge_cells('E33:N33')
    sheet.merge_cells('O33:P33')
    sheet.merge_cells('Q33:R33')
    sheet.merge_cells('S33:T33')

    sheet.merge_cells('A34:T34')

    sheet.merge_cells('A35:E35')
    sheet.merge_cells('F35:L35')
    sheet.merge_cells('M35:T35')

    sheet.merge_cells('A36:E36')
    sheet.merge_cells('F36:L36')
    sheet.merge_cells('M36:T36')

    # Alignments
    center_alignment = Alignment(vertical='center', horizontal='center')
    left_alignment = Alignment(vertical='center', horizontal='left')
    ver_align = Alignment(vertical='center')
    sheet['A1'].alignment = center_alignment

    for cell in sheet['2']:
        cell.alignment = ver_align

    for cell in sheet['3']:
        cell.alignment = left_alignment

    for cell in sheet['4']:
        cell.alignment = ver_align

    for cell in sheet['5']:
        cell.alignment = ver_align

    for cell in sheet['6']:
        cell.alignment = ver_align

    for cell in sheet['7']:
        cell.alignment = ver_align

    for cell in sheet['8']:
        cell.alignment = left_alignment
    
    for cell in sheet['9']:
        cell.alignment = center_alignment

    for cell in sheet['10']:
        cell.alignment = ver_align

    for cell in sheet['11']:
        cell.alignment = center_alignment

    for cell in sheet['12']:
        cell.alignment = center_alignment

    for cell in sheet['13']:
        cell.alignment = ver_align

    for cell in sheet['14']:
        cell.alignment = center_alignment

    for cell in sheet['15']:
        cell.alignment = center_alignment

    for cell in sheet['16']:
        cell.alignment = ver_align

    for cell in sheet['17']:
        cell.alignment = center_alignment
    
    for cell in sheet['18']:
        cell.alignment = center_alignment

    for cell in sheet['19']:
        cell.alignment = center_alignment

    for cell in sheet['20']:
        cell.alignment = center_alignment

    for cell in sheet['21']:
        cell.alignment = center_alignment

    for cell in sheet['22']:
        cell.alignment = center_alignment

    for cell in sheet['23']:
        cell.alignment = center_alignment

    for cell in sheet['25']:
        cell.alignment = left_alignment

    for cell in sheet['26']:
        cell.alignment = left_alignment

    for cell in sheet['28']:
        cell.alignment = left_alignment

    for cell in sheet['29']:
        cell.alignment = left_alignment

    for cell in sheet['31']:
        cell.alignment = ver_align

    for cell in sheet['32']:
        cell.alignment = center_alignment

    for cell in sheet['33']:
        cell.alignment = center_alignment

    for cell in sheet['34']:
        cell.alignment = center_alignment

    for cell in sheet['35']:
        cell.alignment = ver_align

    for cell in sheet['36']:
        cell.alignment = ver_align



    # Height
    sheet.row_dimensions[1].height = 50
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[4].height = 20
    sheet.row_dimensions[5].height = 20
    sheet.row_dimensions[6].height = 20
    sheet.row_dimensions[7].height = 20
    sheet.row_dimensions[8].height = 20
    sheet.row_dimensions[9].height = 15
    sheet.row_dimensions[10].height = 23
    sheet.row_dimensions[11].height = 23
    sheet.row_dimensions[13].height = 23
    sheet.row_dimensions[14].height = 20
    sheet.row_dimensions[15].height = 20
    sheet.row_dimensions[16].height = 23
    sheet.row_dimensions[17].height = 20
    sheet.row_dimensions[18].height = 20
    sheet.row_dimensions[19].height = 20
    sheet.row_dimensions[20].height = 20
    sheet.row_dimensions[21].height = 20
    sheet.row_dimensions[22].height = 20
    sheet.row_dimensions[23].height = 20
    sheet.row_dimensions[24].height = 55
    sheet.row_dimensions[25].height = 23
    sheet.row_dimensions[26].height = 20
    sheet.row_dimensions[27].height = 55
    sheet.row_dimensions[28].height = 23
    sheet.row_dimensions[29].height = 20
    sheet.row_dimensions[30].height = 55
    sheet.row_dimensions[31].height = 23
    sheet.row_dimensions[32].height = 23
    sheet.row_dimensions[33].height = 20
    sheet.row_dimensions[34].height = 30
    sheet.row_dimensions[35].height = 55
    sheet.row_dimensions[36].height = 50

    # Font Size
    a1_font = Font(size=18, bold=True)
    sheet['A1'].font = a1_font

    bold_font = Font(bold=True)

    for cell in sheet['2']:
        cell.font = bold_font

    for cell in sheet['3']:
        cell.font = bold_font

    for cell in sheet['4']:
        cell.font = bold_font

    for cell in sheet['5']:
        cell.font = bold_font

    for cell in sheet['6']:
        cell.font = bold_font

    for cell in sheet['7']:
        cell.font = bold_font

    for cell in sheet['8']:
        cell.font = bold_font
    
    for cell in sheet['9']:
        cell.font = bold_font

    for cell in sheet['10']:
        cell.font = bold_font

    for cell in sheet['11']:
        cell.font = bold_font

    for cell in sheet['12']:
        cell.font = bold_font

    for cell in sheet['13']:
        cell.font = bold_font

    for cell in sheet['14']:
        cell.font = bold_font

    for cell in sheet['15']:
        cell.font = bold_font

    for cell in sheet['16']:
        cell.font = bold_font

    for cell in sheet['17']:
        cell.font = bold_font

    for cell in sheet['18']:
        cell.font = bold_font

    for cell in sheet['19']:
        cell.font = bold_font
    
    for cell in sheet['20']:
        cell.font = bold_font

    for cell in sheet['21']:
        cell.font = bold_font

    for cell in sheet['22']:
        cell.font = bold_font

    for cell in sheet['23']:
        cell.font = bold_font

    for cell in sheet['24']:
        cell.font = bold_font

    for cell in sheet['25']:
        cell.font = bold_font

    for cell in sheet['26']:
        cell.font = bold_font

    for cell in sheet['27']:
        cell.font = bold_font

    for cell in sheet['28']:
        cell.font = bold_font

    for cell in sheet['29']:
        cell.font = bold_font

    for cell in sheet['30']:
        cell.font = bold_font

    for cell in sheet['31']:
        cell.font = bold_font

    for cell in sheet['32']:
        cell.font = bold_font

    for cell in sheet['33']:
        cell.font = bold_font

    for cell in sheet['34']:
        cell.font = bold_font

    for cell in sheet['35']:
        cell.font = bold_font

    for cell in sheet['36']:
        cell.font = bold_font


    # Rows
    sheet['A1'] = 'NON CONFORMANCE & RCA - CAPA REPORT'

    sheet['A2'] = 'Site / Location :'
    sheet['D2'] = data.site
    sheet['L2'] = 'NCR No :'
    sheet['O2'] = data.ncr_number

    sheet['A3'] = 'Customer :'
    sheet['D3'] =  data.customer
    sheet['L3'] = 'NCR Date :'
    sheet['O3'] = data.ncr_date

    sheet['A4'] = 'Product Description :'
    sheet['D4'] = data.product_description
    sheet['L4'] = 'Issued by :'
    sheet['O4'] = data.issued_by

    sheet['A5'] = 'Product Part No :'
    sheet['D5'] = data.product_part_number
    sheet['G5'] = 'Product rev :'
    sheet['I5'] = data.product_rev
    sheet['L5'] = 'Issued to :'
    sheet['O5'] = data.issued_to


    sheet['A6'] = 'Cyient FG No :'
    sheet['D6'] = data.cyient_fg_number
    sheet['L6'] = 'Lot # :'
    sheet['O6'] = data.lot


    sheet['A7'] = 'Issue reported stage :'
    sheet['D7'] = data.issue_reported_stage
    sheet['L7'] = 'Lot & defect status :'
    sheet['O7'] = data.lot_and_defect_status


    sheet['A8'] = 'PO Details'
    sheet['D8'] = data.po_details
    sheet['L8'] = 'NCR Qty'
    sheet['O8'] = data.ncr_quantity


    sheet['A9'] = 'Note :- L = Lot size    S = Sample size     D = Defect quantity     RCA - Root cause Analysis'


    sheet['A10'] = 'PROBLEM DESCRIPTION'


    sheet['A11'] = 'Product serial No.'
    sheet['B11'] = 'Description of Non conformance'
    sheet['N11'] = 'Circuit reference/ Ref designator'
    sheet['R11'] = 'Number of defects'


    sheet['A12'] = data.serial_number
    sheet['B12'] = ''
    sheet['N12'] = data.circuit_reference_or_ref_designator
    sheet['R12'] = data.number_of_defects


    sheet['A13'] = 'CONTAINMENT ACTION : (Containment action to be in place within one working shift)'
    sheet['N13'] = 'Resp.'
    sheet['R13'] = 'Date'

    sheet['A14'] = data.containment_action
    sheet['N14'] = data.containment_resp
    sheet['R14'] = data.containment_date

    sheet['A15'] = ''
    sheet['N15'] = ''
    sheet['R15'] = ''

    sheet['A16'] = 'ROOT CAUSE ANALYSIS : (RCA to be completed within two working shifts)'

    sheet['A17'] = ''
    sheet['B17'] = ''
    sheet['M17'] = 'MAN'
    sheet['O17'] = 'METHOD'
    sheet['Q17'] = 'MATERIAL'
    sheet['S17'] = 'MACHINE'

    # why 1
    sheet['A18'] = 'Why'
    sheet['B18'] = data.why1
    sheet['M18'] = data.why1_man
    sheet['O18'] = data.why1_method
    sheet['Q18'] = data.why1_material
    sheet['S18'] = data.why1_machine


    # why 2
    sheet['A19'] = 'Why'
    sheet['B19'] = data.why2
    sheet['M19'] = data.why2_man
    sheet['O19'] = data.why2_method
    sheet['Q19'] = data.why2_material
    sheet['S19'] = data.why2_machine


    # why 3
    sheet['A20'] = 'Why'
    sheet['B20'] = data.why3
    sheet['M20'] = data.why3_man
    sheet['O20'] = data.why3_method
    sheet['Q20'] = data.why3_material
    sheet['S20'] = data.why3_machine


    # why 4
    sheet['A21'] = 'Why'
    sheet['B21'] = data.why4
    sheet['M21'] = data.why4_man
    sheet['O21'] = data.why4_method
    sheet['Q21'] = data.why4_material
    sheet['S21'] = data.why4_machine


    # why 5
    sheet['A22'] = 'Why'
    sheet['B22'] = data.why5
    sheet['M22'] = data.why5_man
    sheet['O22'] = data.why5_method
    sheet['Q22'] = data.why5_material
    sheet['S22'] = data.why5_machine


    # How
    sheet['A23'] = 'How'
    sheet['B23'] = data.how
    sheet['M23'] = 'List all probabilities & encircle most likely contribution'


    note1 = sheet['A24']
    note1.value = 'Note: \n 1. Why why analysis must have proper link between first why to next why while identifying the final root cause of problem. \n 2. Root cause analysis shall consider the causes with respect to Human Factors also. '
    note1.alignment = Alignment(wrap_text=True, vertical='center')

    sheet['A25'] = 'CORRECTIVE  ACTION : (CA to be deployed within three working shifts)'
    sheet['N25'] = 'Resp.'
    sheet['R25'] = 'Date'

    sheet['A26'] = data.corrective_action
    sheet['N26'] = data.corrective_resp
    sheet['R26'] = data.corrective_date

    note2 = sheet['A27']
    note2.value = 'Note: \n 1. Actions must match with the final root cause identified during root cause analysis. Also have either refresher training or document level corrections. \n 2. Action shall cover the improvement proposed to fix the causes with respect to Human Factors also. '
    note2.alignment = Alignment(wrap_text=True, vertical='center')

    sheet['A28'] = 'PREVENTIVE  ACTION/ HORIZONTAL DEPLOYMENT:  (PA to be deployed within four working shifts)'
    sheet['N28'] = 'Resp.'
    sheet['R28'] = 'Date'

    sheet['A29'] = data.preventive_action
    sheet['N29'] = data.preventive_resp
    sheet['R29'] = data.preventive_date

    note2 = sheet['A30']
    note2.value = 'Note: \n 1. Actions must match with the final root cause identified during root cause analysis. Also have either refresher training or document level corrections. \n 2. Action shall cover the improvement proposed to fix the causes with respect to Human Factors also. '
    note2.alignment = Alignment(wrap_text=True, vertical='center')

    sheet['A31'] = 'VERIFICATION FOR EFFECTIVENESS  (To be filled by originating function)'

    sheet['A32'] = 'WO No.'
    sheet['C32'] = 'Batch No.'
    sheet['E32'] = 'Verification details'
    sheet['O32'] = 'Status'
    sheet['Q32'] = 'Resp.'
    sheet['S32'] = 'Date'

    sheet['A33'] = data.wo_no
    sheet['C33'] = data.batch_no
    sheet['E33'] = data.verification_details
    sheet['O33'] = data.verification_status
    sheet['Q33'] = data.verification_resp
    sheet['S33'] = data.varifiction_date

    sheet['A34'] = f"Status: {data.status}"

    sheet['A35'] = 'Prepared by: (Action initiator)'
    sheet['F35'] = 'Approved by: (Manager - Quality)'
    sheet['M35'] = 'Approved by:'

    sheet['A36'] = 'Signature & Date:'
    sheet['F36'] = 'Signature & Date:'
    sheet['M36'] = 'Signature & Date:'

    # Image Start

    # Load and resize images
    max_width = 370
    max_height = 370

    img1 = PILImage.open(data.image1)
    img1.thumbnail((max_width, max_height))

    img2 = PILImage.open(data.image2)
    img2.thumbnail((max_width, max_height))

    spacer_width = 20  # Adjust the width of the spacer image as needed
    spacer = PILImage.new('RGB', (spacer_width, max(img1.height, img2.height)), (255, 255, 255))  # Create a white spacer image

    # Combine the images with the spacer
    composite_image = PILImage.new('RGB', (img1.width + img2.width + spacer_width, max(img1.height, img2.height)))
    composite_image.paste(img1, (0, 0))
    composite_image.paste(spacer, (img1.width, 0))
    composite_image.paste(img2, (img1.width + spacer_width, 0))

    # Save the composite image
    composite_image.save('composite_image.jpg')

    # Add the image to cell B12
    img = Image('composite_image.jpg')
    sheet.add_image(img, 'B12')

    # Adjust row height for the second row based on image height
    #sheet.row_dimensions[12].height = img1.height
    sheet.row_dimensions[12].height = 200

    # Image End

    # Logo Start

    image_file_path = 'static/images/CyientDLM.jpg'

    # Check if the image file exists
    if os.path.exists(image_file_path):
        # Add the image to the first row (row 1)
        img3 = ExcelImage(image_file_path)
        sheet.add_image(img3, 'A1')

        # Set row height for the first row based on image height
        sheet.row_dimensions[1].height = img3.height
    else:
        # Handle the case where the image file does not exist
        print(f"Image file '{image_file_path}' does not exist.")

    # Logo End

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="NCR.xlsx"'

    workbook.save(response)

    return response






import csv

def import_data_from_csv(request):
    with open(r'C:\Users\Administrator\Downloads\dashboard_allserial.csv', mode='r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            serial_number = SerialNumber(
                project=row['project'],
                po_number=row['po_number'],
                ncr_number=row['ncr_number'],
                date=row['date'],
                customer_part_number=row['customer_part_number'],
                cyient_part_number=row['cyient_part_number'],
                quantity=row['quantity'],
                serial_number=row['serial_number']
            )
            serial_number.save()

